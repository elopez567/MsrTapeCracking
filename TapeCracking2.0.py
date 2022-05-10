# library
import re
import pandas as pd
import openpyxl as xl
from os import listdir
from os.path import isfile, join

# from win32com.client import Dispatch


class CrackTape:
    def __init__(self):

        self.get_paths()
        cut_date = self.dlq_cut_date()

        print(f"\n{bcolors.OKCYAN}{bcolors.BOLD}Running...")

        self.dicts_and_lists()
        self.load_excelwb()

        self.new_dict_keys = []  # Blank list for our fields
        self.new_dict_val = []  # Blank list for seller's column numbers

        self.match_fields()

        self.new_dict = dict(zip(self.new_dict_keys, self.new_dict_val))

        self.transfer_data()
        self.missing_data()

    def dicts_and_lists(self):
        self.our_fields = {'Loan_NO': 1, "Amort_Term": 2, 'App_Value': 3, 'Balance': 4, "Current_Payment": 8,
                           "Curr_Rate": 9, 'Document': 11, 'Fico_Score': 12, 'First_Pay_Date': 13, 'LPI_Date': 25,
                           'Occupancy': 29, 'Orig_Amt': 30, 'Orig_Date': 32, 'Orig_LTV': 33, 'Product_Type': 42,
                           'Prop_Type': 43, 'Purpose': 44, 'State': 48, 'Zip': 50, 'City': 53, 'Foreclosure': 55,
                           'Modified': 57, 'Bankruptcy': 58, 'Mod Date': 59, 'Gfee': 64, 'Net Sfee': 65, 'Escrow': 66}

        self.searchers = {'Loan_NO': '\Aloan(_|\s)?(No|ID|Nu\w+)', "Amort_Term": 'Orig.(Loan.)?Term(-Calc)?',
                          'App_Value': '(Or\w+.)?ap\w+(.Va\w+)?', 'Balance': '(^UPB|Cur\w+.(Loan.)?Ba\w+)',
                          "Current_Payment": '(^P&I|^PI)', "Curr_Rate": '^In\w+.Rate', 'Document': r'\ADoc(.+)?',
                          'Fico_Score': '\AFICO(.SCORE)?$', 'First_Pay_Date': '(First(.+)?Date|\AFPD)',
                          'LPI_Date': 'Next(.+)Date',
                          'Occupancy': '^Oc\w+', 'Orig_Amt': 'Ori\w+(\s)?(Loan)?(_|\s)?Ba\w+',
                          'Orig_Date': '(Origi(nation)$|Ori(.+)?Date)', 'Orig_LTV': '(\w+)?(\s|_)?LTV(\s|_)?(\w+)?',
                          'Product_Type': '(\AType|\AProd(\w\w\w)?(.)?(\w+)?)|\ALoan.(Prod.+)?Type',
                          'Prop_Type': '\AProp(.+)?',
                          'Purpose': '(loan.)?Purpose', 'State': 'State', 'Zip': 'Zip(\s|_)?(\w+)?', 'City': 'city',
                          'Foreclosure': 'FC', 'Modified': '\AMod(\w+)?(.)?(Loan)?', 'Bankruptcy': 'BK',
                          'Mod Date': '(.+)?mo\w+(.)date', 'Gfee': '\AG(.+)Fee', 'Net Sfee': '\ANet.Ser.+Fee',
                          'Escrow': '(\AT&I|\ATI)'}

        self.search_keys = list(self.searchers.keys())
        self.search_values = list(self.searchers.values())

    def load_excelwb(self):
        # Opens workbooks and pulls seller's fields
        self.seller_wb = xl.load_workbook(self.seller_tape, data_only=True)
        self.sws = self.seller_wb.worksheets[0]  # Seller worksheet
        self.our_wb = xl.load_workbook(self.blank_tape, data_only=True)
        self.ows = self.our_wb.worksheets[0]  # Our blank worksheet

        seller_data = pd.read_excel(io=self.seller_tape).columns

        x = 1
        self.seller_index = {}
        # # Creates seller tape fields:index dictionary
        for field in seller_data:
            self.seller_index[field] = x
            x += 1

        del seller_data

        # list of Seller's fields
        self.seller_fields = list(self.seller_index.keys())

    def get_paths(self):
        self.my_path = r'M:\Capital Markets\Users\Emmanuel Lopez\Python\Python Tape Cracking\SELLER TAPE'
        self.out_path = r'M:\Capital Markets\Users\Emmanuel Lopez\Python\Python Tape Cracking\OUTPUT'
        self.seller_tape = "".join([self.my_path, '\\',
                                    [f for f in listdir(self.my_path) if isfile(join(self.my_path, f))][0]])
        self.blank_tape = "".join([self.out_path, '\\', 'LEAVE_BLANK_2018mmdd_PNMACFile.xlsx'])

        # USED FOR TO CREATE CUTDATE FOR DLQ STATUS PAGE LATER

    def dlq_cut_date(self):
        cut_date = str(input(f"\n{bcolors.WARNING}{bcolors.BOLD}Please enter cutoff date as mm/dd/yyyy: "))
        month, day, year = cut_date[:2], '1', cut_date[-4:]
        if month == '12':
            month = '1'
            year = str(int(year) + 1)
        else:
            month = str(int(month) + 1)
        return f'{month}/{day}/{year}'

    def get_key(self, val):
        for key, value in self.searchers.items():
            if val == value:
                self.new_dict_keys.append(key)

    def match_fields(self):
        print(f"{bcolors.OKCYAN}{bcolors.BOLD}Matching Fields...")
        # Uses Regex to match our fields to seller tape, if there is a match, the column number of the seller's field is
        # appended to NewDictValues
        for Key in self.search_keys:
            for x in range(len(self.seller_fields)):
                regex = self.searchers.get(Key)
                p = re.compile(regex, re.IGNORECASE)
                m = p.match(self.seller_fields[x])
                if m:
                    self.get_key(regex)
                    self.new_dict_val.append(self.seller_index[self.seller_fields[x]])
                    break

    def transfer_data(self):
        # pulls seller tape index and our index and integrates excel to copy data over
        mr = self.sws.max_row
        print(f"{bcolors.OKCYAN}{bcolors.BOLD}Transferring Data...")
        for key, val in self.new_dict.items():
            for i in range(2, mr + 1):
                c = self.sws.cell(row=i, column=val)  # pulls seller's data
                our_col = self.our_fields[key]  # our field index
                self.oc = self.ows.cell(row=i, column=int(our_col))
                self.oc.value = c.value  # transfer of data
                self.conditionals(key)
        self.our_wb.save(self.out_path + '\\Finished_2018mmdd_PNMACFile.xlsx')

    def conditionals(self, key):
        if self.our_fields[key] in [13, 23, 32, 59]:
            self.oc.number_format = 'm/d/yyyy'
        if self.our_fields[key] in [55, 57, 58]:
            if self.oc.value is None:
                self.oc.value = 0
        if self.our_fields[key] in [9, 33, 64, 65]:
            self.oc.number_format = '0.####'
            if self.oc.value is None:
                pass
            elif float(self.oc.value) > 1:
                self.oc.value = float(self.oc.value) / 100
            elif float(self.oc.value) < 0.01:
                self.oc.value = float(self.oc.value) * 100

    def missing_data(self):
        # Checks and logs missing matches
        missing = []
        for key in self.our_fields.keys():
            if key not in self.new_dict_keys:
                missing.append(key)
        return print(f"{bcolors.FAIL}{bcolors.BOLD}\nMissing values: {str(missing)}")


class bcolors:
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'


CrackTape()
