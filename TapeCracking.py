# library
import re
import pandas as pd
import openpyxl as xl
from os import listdir
from os.path import isfile, join
from win32com.client import Dispatch


class bcolors:
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'


# Path to SELLER TAPE/OUTPUT folder & file
mypath = r'M:\Capital Markets\Users\Emmanuel Lopez\Python\Python Tape Cracking\SELLER TAPE'
OutPath = r'M:\Capital Markets\Users\Emmanuel Lopez\Python\Python Tape Cracking\OUTPUT'
SellerTapeFile = [f for f in listdir(mypath) if isfile(join(mypath, f))]
SellerTapePath = mypath + '\\' + SellerTapeFile[0]
BlankTapePath = OutPath + '\\' + 'LEAVE_BLANK_2018mmdd_PNMACFile.xlsx'

cutdate = str(input(f"\n{bcolors.WARNING}{bcolors.BOLD}Please enter cutoff date as mm/dd/yyyy: "))

print(f"\n{bcolors.OKCYAN}{bcolors.BOLD}Running...")

# Dictionary with our tape fields/corresponding columns numbers for later indexing
OurTape = {'Loan_NO': 1, "Amort_Term": 2, 'App_Value': 3, 'Balance': 4, "Current_Payment": 8, "Curr_Rate": 9,
           'Document': 11, 'Fico_Score': 12, 'First_Pay_Date': 13, 'LPI_Date': 25, 'Occupancy': 29, 'Orig_Amt': 30,
           'Orig_Date': 32, 'Orig_LTV': 33, 'Product_Type': 42, 'Prop_Type': 43, 'Purpose': 44, 'State': 48, 'Zip': 50,
           'City': 53, 'Foreclosure': 55, 'Modified': 57, 'Bankruptcy': 58, 'Mod Date': 59, 'Gfee': 64, 'Net Sfee': 65,
           'Escrow': 66}

# Dictionary for our fields & corresponding regex
Searchers = {'Loan_NO': '\Aloan(_|\s)?(No|ID|Nu\w+)', "Amort_Term": 'Orig.(Loan.)?Term(-Calc)?',
             'App_Value': '(Or\w+.)?ap\w+(.Va\w+)?', 'Balance': '(^UPB|Cur\w+.(Loan.)?Ba\w+)',
             "Current_Payment": '(^P&I|^PI)', "Curr_Rate": '^In\w+.Rate', 'Document': r'\ADoc(.+)?',
             'Fico_Score': '\AFICO(.SCORE)?$', 'First_Pay_Date': '(First(.+)?Date|\AFPD)', 'LPI_Date': 'Next(.+)Date',
             'Occupancy': '^Oc\w+', 'Orig_Amt': 'Ori\w+(\s)?(Loan)?(_|\s)?Ba\w+',
             'Orig_Date': '(Origi(nation)$|Ori(.+)?Date)', 'Orig_LTV': '(\w+)?(\s|_)?LTV(\s|_)?(\w+)?',
             'Product_Type': '(\AType|\AProd(\w\w\w)?(.)?(\w+)?)|\ALoan.(Prod.+)?Type', 'Prop_Type': '\AProp(.+)?',
             'Purpose': '(loan.)?Purpose', 'State': 'State', 'Zip': 'Zip(\s|_)?(\w+)?', 'City': 'city',
             'Foreclosure': 'FC', 'Modified': '\AMod(\w+)?(.)?(Loan)?', 'Bankruptcy': 'BK',
             'Mod Date': '(.+)?mo\w+(.)date', 'Gfee': '\AG(.+)Fee', 'Net Sfee': '\ANet.Ser.+Fee',
             'Escrow': '(\AT&I|\ATI)'}

# List for Searchers Keys/Values for regex and indexing use later
SearchVal = list(Searchers.values())
SearchKey = list(Searchers.keys())


### This Section will scrape seller's tape fields and create a dictionary with it's corresponding column number ###
# Sets workbook/sheets to variables
SellerWB = xl.load_workbook(SellerTapePath, data_only=True)
SWS = SellerWB.worksheets[0]
OurWB = xl.load_workbook(BlankTapePath, data_only=True)
OWS = OurWB.worksheets[0]

# Reads Seller's Tape fields
df = pd.read_excel(io=SellerTapePath).columns

SellTape = {}
x = 1

# Creates seller tape fields:index dictionary
for Fields in df:
    SellTape[Fields] = x
    x += 1

# list of Seller's fields
STlist = list(SellTape.keys())

NewDictKeys = []  # Blank list for our fields
NewDictValues = []  # Blank list for seller's column numbers

### FUNCTIONS ###
# function pulls our tape fields from Searchers dict and adds to NewDictKeys list
def get_key(val):
    for key, value in Searchers.items():
        if val == value:
            NewDictKeys.append(key)

# Conditional function checks our fields for formatting purposes
def conditionals(key):
    if (OurTape[key] == 13) or (OurTape[key] == 23) or (OurTape[key] == 32) or (OurTape[key] == 59):
        OC.number_format = 'm/d/yyyy'
    if (OurTape[key] == 55) or (OurTape[key] == 57) or (OurTape[key] == 58):
        if OC.value is None:
            OC.value = 0
    if (OurTape[key] == 9) or (OurTape[key] == 33):
        OC.number_format = '0.####'
        if OC.value is None or type(OC.value) == str:
            pass
        elif float(OC.value) > 1:
            OC.value = float(OC.value) / 100
    if (OurTape[key] == 64) or (OurTape[key] == 65):
        OC.number_format = '0.####'
        if OC.value is None:
            pass
        elif float(OC.value) < 0.01:
            OC.value = float(OC.value) * 100


# USED FOR TO CREATE CUTDATE FOR DLQ STATUS PAGE LATER
month = cutdate[:2]
day = '1'
year = cutdate[-4:]
if month == '12':
    month = '1'
    year = str(int(year)+1)
else:
    month = str(int(month)+1)
DLQCutDate = f'{month}/{day}/{year}'

### MATCHING ###
print(f"{bcolors.OKCYAN}{bcolors.BOLD}Matching Fields...")
# Uses Regex to match our fields to seller tape, if there is a match, the column number of the seller's field is
# appended to NewDictValues
for Key in SearchKey:
    for x in range(len(STlist)):
        RegexVal = Searchers.get(Key)
        p = re.compile(RegexVal, re.IGNORECASE)
        m = p.match(STlist[x])
        if m:
            get_key(RegexVal)
            NewDictValues.append(SellTape[STlist[x]])
            break

Missing = []

# Checks and logs missing matches
for key in OurTape.keys():
    if key not in NewDictKeys:
        Missing.append(key)

# Creates new Dict that pairs our fields with column number of the seller's fields
NewDict = dict(zip(NewDictKeys, NewDictValues))

# records max rows
mr = SWS.max_row

### this section copies data from seller tape over to our tape ###
print(f"{bcolors.OKCYAN}{bcolors.BOLD}Transferring Data...")

# pulls seller tape index and our index and integrates excel to copy data over
for key, val in NewDict.items():
    for i in range(2, mr + 1):
        c = SWS.cell(row=i, column=val)  # pulls seller's data
        ourCol = OurTape[key]  # our field index
        OC = OWS.cell(row=i, column=int(ourCol))
        OC.value = c.value  # transfer of data
        conditionals(key)

# saves first data transfer
OurWB.save(OutPath + '\\Finished_2018mmdd_PNMACFile.xlsx')

###  VBA IN PYTHON - COPY DUPLICATES   ###
TapeFile = OutPath + '\\Finished_2018mmdd_PNMACFile.xlsx'
wkbk1 = TapeFile
excel = Dispatch("Excel.Application")
excel.Visible = 1
source = excel.Workbooks.Open(wkbk1)

# MAIN PAGE
excel.Worksheets(1).Select()
excel.Range(f"A2:A{mr}").Select()
excel.Selection.Copy()
excel.Range(f"W2:W{mr}").PasteSpecial(Paste=-4163)

excel.Range(f"D2:D{mr}").Select()
excel.Selection.Copy()
excel.Range(f"E2:E{mr}").PasteSpecial(Paste=-4163)

excel.Range(f"I2:I{mr}").Select()
excel.Selection.Copy()
excel.Range(f"R2:R{mr}").PasteSpecial(Paste=-4163)

excel.Range(f"AG2:AG{mr}").Select()
excel.Selection.Copy()
excel.Range(f"AE2:AE{mr}").PasteSpecial(Paste=-4163)

excel.Range("X2").formula = '=edate(Y2,-1)'
excel.Range("X2").Select()
excel.Selection.Copy()
excel.Range(f"X3:X{mr}").PasteSpecial(Paste=-4123)

excel.Range("V2").value = '1'
excel.Range("V2").Select()
excel.Selection.Copy()
excel.Range(f"V3:V{mr}").PasteSpecial(Paste=12)

excel.Range("AN2").value = '2'
excel.Range("AN2").Select()
excel.Selection.Copy()
excel.Range(f"AN3:AN{mr}").PasteSpecial(Paste=12)

excel.Range("BB2").value = '0'
excel.Range("BB2").Select()
excel.Selection.Copy()
excel.Range(f"BB3:BB{mr}").PasteSpecial(Paste=12)

excel.Range("BD2").value = '0'
excel.Range("BD2").Select()
excel.Selection.Copy()
excel.Range(f"BD3:BD{mr}").PasteSpecial(Paste=12)

if cutdate[0] == '0':
    excel.Range("AY2").value = cutdate[1:]
else:
    excel.Range("AY2").value = cutdate
excel.Range("AY2").Select()
excel.Selection.Copy()
excel.Range(f"AY3:AY{mr}").PasteSpecial(Paste=12)
excel.Range('AY:AY').NumberFormat = "m/d/yyyy"

# LOAN TYPE PAGE
excel.Worksheets(1).Select()
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(3).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)

# PROP TYPE PAGE
excel.Worksheets(1).Select()
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(4).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("AQ:AQ").Select()
excel.Selection.Copy()
excel.Worksheets(4).Select()
excel.Range("B:B").PasteSpecial(Paste=-4163)

# BPO PAGE
excel.Worksheets(1).Select()
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(6).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)

# ZIP PAGE
excel.Worksheets(1).Select()
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(9).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("AX:AX").Select()
excel.Selection.Copy()
excel.Worksheets(9).Select()
excel.Range("B:B").PasteSpecial(Paste=-4163)

excel.Range("D2").formula = '=TEXT(B2,"00000")'
excel.Range("D2").Select()
excel.Selection.Copy()
excel.Range(f"D3:D{mr}").PasteSpecial(Paste=-4123)

# TERM CHECK PAGE
excel.Worksheets(1).Select()
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(10).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("B:B").Select()
excel.Selection.Copy()
excel.Worksheets(10).Select()
excel.Range("B:B").PasteSpecial(Paste=-4163)
excel.Range("B1").value = 'TERM'
excel.Range("C:C").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("M:M").Select()
excel.Selection.Copy()
excel.Worksheets(10).Select()
excel.Range("D:D").PasteSpecial(Paste=12)

excel.Worksheets(1).Select()
excel.Range("AY:AY").Select()
excel.Selection.Copy()
excel.Worksheets(10).Select()
excel.Range("F:F").PasteSpecial(Paste=12)

excel.Range("I2").formula = '=DATEDIF(D2, E2, "M") + 1'
excel.Range("J2").formula = '=I2 + (C2 - B2)'
excel.Range("K2").formula = '=I2 - B2'
excel.Range("L2").formula = '=IF( DAY(D2)*DAY(E2) > 1, TRUE, FALSE)'
excel.Range("M2").formula = '=I2'
excel.Range("N2").formula = '=J2'
excel.Range("O2").formula = '=DATEDIF(D2, F2, "M") + 1'
excel.Range("P2").formula = '= M2- O2'
excel.Range("Q2").formula = '=ROUNDUP((60 - P2) / 12, 0) * 12'
excel.Range("R2").formula = '=M2 + Q2'
excel.Range("S2").formula = '=N2 + Q2'
excel.Range("I2:S2").Select()
excel.Selection.Copy()
excel.Range(f"I3:S{mr}").PasteSpecial(Paste=-4123)

# LTV CHECK PAGE
excel.Worksheets(1).Select()
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(11).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("AG:AG").Select()
excel.Selection.Copy()
excel.Worksheets(11).Select()
excel.Range("B:B").PasteSpecial(Paste=-4163)
excel.Range("C:C").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("AD:AD").Select()
excel.Selection.Copy()
excel.Worksheets(11).Select()
excel.Range("D:D").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("C:C").Select()
excel.Selection.Copy()
excel.Worksheets(11).Select()
excel.Range("E:E").PasteSpecial(Paste=-4163)

excel.Range("G2").formula = '=MAX(B2, D2/MIN(E2,F2))'
excel.Range("H2").formula = '=(C2-B2)+G2'
excel.Range("G2:H2").Select()
excel.Selection.Copy()
excel.Range(f"G3:H{mr}").PasteSpecial(Paste=-4123)

# DLQ STATUS PAGE
excel.Worksheets(1).Select()
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(12).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)

excel.Range(f"C2:C{mr}").value = DLQCutDate
excel.Range('C:C').NumberFormat = "m/d/yyyy"

excel.Worksheets(1).Select()
excel.Range("BC:BF").Select()
excel.Selection.Copy()
excel.Worksheets(12).Select()
excel.Range("F:I").PasteSpecial(Paste=-4163)

excel.Worksheets(1).Select()
excel.Range("X:X").Select()
excel.Selection.Copy()
excel.Range("X:X").PasteSpecial(Paste=-4163)
excel.Range('X:X').NumberFormat = "m/d/yyyy"

excel.Range("X:X").Select()
excel.Selection.Copy()
excel.Worksheets(12).Select()
excel.Range("B:B").PasteSpecial(Paste=12)

excel.Range("D2").formula = '=DATEDIF(B2, C2, "M")'
excel.Range("D2").Select()
excel.Selection.Copy()
excel.Range(f"D3:D{mr}").PasteSpecial(Paste=-4123)

excel.Range("E2").formula = '=IF(F2=1,5,IFERROR(IF(D2>3,4,VLOOKUP(D2,$L$4:$M$8,2,0)),1))'
excel.Range("E2").Select()
excel.Selection.Copy()
excel.Range(f"E3:E{mr}").PasteSpecial(Paste=-4123)

excel.Worksheets(1).Select()
excel.Range(f"Y2:Y{mr}").Select()
excel.Selection.Clear()

# PAY HISTORY
excel.Range("A:A").Select()
excel.Selection.Copy()
excel.Worksheets(13).Select()
excel.Range("A:A").PasteSpecial(Paste=-4163)
excel.Worksheets(1).Select()

source.Save()
print(f"{bcolors.OKGREEN}{bcolors.BOLD}\nDone!")
print(f"{bcolors.FAIL}{bcolors.BOLD}\nMissing values: {str(Missing)}")
